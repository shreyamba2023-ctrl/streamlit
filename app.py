import streamlit as st
import pandas as pd
from pathlib import Path
import json
import logging
import re
from tqdm import tqdm
import pdfplumber, docx
import instructor
from pydantic import BaseModel, Field
from typing import Dict, Optional, List
from docx2python import docx2python
from functools import lru_cache
import base64
from datetime import datetime
from openai import OpenAI
from docx import Document
import country_converter as coco
import os
from dotenv import load_dotenv
import google.generativeai as genai
import math
import faiss
import gc
import numpy as np
import time
from sentence_transformers import SentenceTransformer
from openpyxl import Workbook, load_workbook

@st.cache_resource
def load_embedder():
    return SentenceTransformer("all-MiniLM-L6-v2")

embedder = load_embedder()

ARTIFACT_DIR = Path("artifacts")
ARTIFACT_DIR.mkdir(exist_ok=True)

FAISS_INDEX_PATH = ARTIFACT_DIR / "rag_index.faiss"
METADATA_PATH = ARTIFACT_DIR / "rag_metadata.json"
EMBED_TEXT_PATH = ARTIFACT_DIR / "rag_embedding_texts.json"

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)
logger = logging.getLogger("talent-search")

if "GEMINI_API_KEY" in st.secrets:
    st.success("API Key found in st.secrets!")
    # Print only the first/last characters to verify it's the correct key
    raw_key = st.secrets["GEMINI_API_KEY"]
    st.write(f"Key preview: {raw_key[:4]}...{raw_key[-4:]}")
else:
    st.error("API Key NOT found in st.secrets. Check your Streamlit Cloud settings."))

MODEL_NAME = "models/gemini-2.5-pro"
GEMINI_BASE_URL = "https://generativelanguage.googleapis.com/v1beta/openai/"

EXCEL_PATH = "candidate_feature_store.xlsx"
OUTPUT_EXCEL = "candidate_feature_store.xlsx"

RESUME_FOLDER = Path(
    "Resumes"
)

class WorkEntry(BaseModel):
    company: str
    job_title: str
    years: float
    sector: str = Field(description="Technology, Healthcare, Financial Services, etc.")
    investment_strategy: Optional[str] = Field(
        description="Fundamental/Discretionary or Quantitative/Systematic"
    )
    summary: str
    is_competitor: bool
    recency_score: float


class EducationEntry(BaseModel):
    university: str
    degree: str
    major: str
    gpa: float
    graduation_year: int
    summary: str
    recency_score: float


# ---------------- Aggregate Feature Blocks ----------------

class EducationFeatures(BaseModel):
    highest_degree: str
    institution_tier: str
    institution_score: float
    gpa_normalized: float
    cfa_level: str
    education_history: List[EducationEntry]

class SkillSignal(BaseModel):
    mentions: int = Field(description="How many times the skill is mentioned or implied")
    years_used: float = Field(description="Estimated total years the skill was used")
    last_used_years_ago: float = Field(
        description="How many years ago the skill was last used (0 = current role)"
    )
    project_count: int = Field(
        description="Number of distinct projects or roles where skill was applied"
    )
    evidence: List[str] = Field(
        description="1–5 short resume snippets supporting the signal"
    )


class WorkFeatures(BaseModel):
    dominant_side: str                  # buyside / sellside / mixed
    buyside_years_last_5: float
    sellside_years_last_5: float
    work_quality_score: float
    work_history: List[WorkEntry]


class TechnicalFeatures(BaseModel):
    financial_modelling: SkillSignal
    python: SkillSignal
    SQL: SkillSignal
    C: SkillSignal
    R: SkillSignal
    machine_learning: SkillSignal
    time_series_modelling: SkillSignal
    deep_learning: SkillSignal
    bloomberg: SkillSignal
    Excel_VBA: SkillSignal
    tableau: SkillSignal
    powerBI: SkillSignal
    additional_skills: str



class DemographicFeatures(BaseModel):
    total_years_experience: float
    primary_geography: str
    extracurriculars: str


# ---------------- Root Profile ----------------

class FeatureProfile(BaseModel):
    name: str
    education: EducationFeatures
    work: WorkFeatures
    technical: TechnicalFeatures
    demographics: DemographicFeatures


def recency_decay(years_ago: float) -> float:
    """
    Exponential decay for recency.
    0 years ago  -> 1.0
    1 year ago   -> ~0.50
    3 years ago  -> ~0.12
    """
    return math.exp(-0.7 * years_ago)


def compute_skill_score(signal) -> float:
    """
    Deterministic skill strength score in range ~[0, 1.2].
    Can later be rescaled if needed.
    """
    if signal.mentions == 0:
        return 0.0

    score = (
        0.35 * math.log1p(signal.mentions) +
        0.35 * min(signal.years_used / 5.0, 1.0) +
        0.20 * recency_decay(signal.last_used_years_ago) +
        0.10 * min(signal.project_count / 5.0, 1.0)
    )

    return round(score, 4)
SKILL_FIELDS = [
    "financial_modelling",
    "python",
    "SQL",
    "C",
    "machine_learning",
    "time_series_modelling",
    "deep_learning",
    "bloomberg",
    "Excel_VBA",
    "tableau",
    "powerBI"
]
def build_skill_columns(technical_features) -> Dict[str, object]:
    """
    Flattens all SkillSignal fields into Excel-friendly columns.
    Adds computed skill_score per skill.
    """

    row = {}

    for skill in SKILL_FIELDS:
        signal = getattr(technical_features, skill)

        score = compute_skill_score(signal)

        prefix = skill.lower()

        row[f"{prefix}_mentions"] = signal.mentions
        row[f"{prefix}_years_used"] = signal.years_used
        row[f"{prefix}_last_used_years_ago"] = signal.last_used_years_ago
        row[f"{prefix}_project_count"] = signal.project_count
        row[f"{prefix}_evidence"] = " | ".join(signal.evidence)
        row[f"{prefix}_skill_score"] = score

    return row


def llm_feature_engineering(raw_resume_text: str) -> FeatureProfile:
    """
    Uses LLM to convert structured resume profile into scored features.
    """

    parsed = client.chat.completions.create(
        model=MODEL_NAME,
        response_model=FeatureProfile,
        temperature=0,
        messages=[
            {
                "role": "system",
                "content": """
You are a senior quantitative recruiter and talent intelligence engine.

Your task:
- Extract every individual work experience and academic degree as separate entries in the provided lists.
- Consider only full-time experience and not internships.
- The investment strategies shall be categorised into fundamental/discretionary, quantitative/systematic. Mark this field as None if the role is not related to any of these.
- Classify the candidates into one of the sectors: Technology, Healthcare, Financial Services, Energy, Industrials, Consumer, Credit, Macro, Equity.
- Use global university awareness (US, Europe, India, Asia) while devising the tier of university. For instance, IITs/NITs/BITS/DTU/NSIT are Tier-1 in India and other government colleges are Tier-2 and rest are tier-3.
- Normalize the GPA of candidates by outputing (GPA scored/maximum GPA in University) in the gpa field. Assign 0 if not mentioned.
- The summaries should have a jist of the work done and any specific achievement during that period.
- Infer buy-side vs sell-side realistically (look for keywords for whom the analysis is being done, for portfolio managers, it is a buy-side role, for direct clients it should be sell-side).
- Estimate only last 5 years of experience for work metrics for identifying the buy-side and sell-side work experience.
- Extract every work experience and academic degree as individual entries.
- Categorize each job into a Sector and an Investment Strategy (Fundamental/Systematic).
- While mentioning the additional skills, avoid adding spoken languages. It should strictly cover coding/financial skills that the candidate has.
- Identify if the firm is a 'Competitor' (major hedge fund like Citadel, D.E. Shaw, Goldman Sachs, JP Morgan).
- For every technical skill, extract objective signals instead of assigning scores.

For each skill provide:
- mentions: integer count of explicit or implicit mentions.
- years_used: estimated total years the skill was used professionally.
- last_used_years_ago: estimate how many years ago the skill was last used (0 if current role).
- project_count: number of distinct projects or roles where the skill was applied.
- evidence: 1–5 literal snippets copied from the resume supporting the signal.

Rules:
- If the skill is not mentioned, set all numeric fields to 0 and evidence = [].
- Do NOT fabricate projects or experience.
- Evidence must be copied verbatim from the resume text.
- Be conservative in estimation when uncertain.
- Numeric outputs must be stable and reasonable.
- Produce stable numeric outputs.

The recency scores should be assigned with respect to current date.
Return strictly structured output.
"""
            },
            {
                "role": "user",
                "content": f"""
Resume Profile:
{raw_resume_text}
"""
            }
        ],
    )

    return parsed

genai.configure(api_key=GEMINI_API_KEY)

GEMINI_BASE_URL = "https://generativelanguage.googleapis.com/v1beta/openai/"
MODEL_NAME = "models/gemini-2.5-pro"

openai_client = OpenAI(
    api_key=GEMINI_API_KEY,
    base_url=GEMINI_BASE_URL
)

client = instructor.from_openai(openai_client)


def extract_text(file: Path) -> str:
    if file.suffix.lower() == ".docx":
        try:
            print(f"📄 Parsing DOCX with docx2python: {file.name}")

            doc = docx2python(str(file))

            blocks = []

            # body_runs contains paragraphs + tables flattened in reading order
            for item in doc.body_runs:
                if isinstance(item, list):
                    # table row or nested structure
                    flat = []
                    def flatten(x):
                        for i in x:
                            if isinstance(i, list):
                                flatten(i)
                            else:
                                txt = str(i).strip()
                                if txt:
                                    flat.append(txt)

                    flatten(item)
                    if flat:
                        blocks.append(" | ".join(flat))
                else:
                    txt = str(item).strip()
                    if txt:
                        blocks.append(txt)

            text = "\n".join(blocks)
            return text.strip()

        except Exception as e:
            print(f"❌ DOCX parsing failed for {file.name}: {e}")
            return ""

    # ================= PDF =================
    elif file.suffix.lower() == ".pdf":
        try:
            print(f"📄 Gemini parsing PDF: {file.name}")

            uploaded_file = genai.upload_file(path=str(file))
            model = genai.GenerativeModel(MODEL_NAME)

            prompt = """
Extract ALL readable text from this resume.
Rules:
- Preserve table rows using pipe separator: Col1 | Col2 | Col3
- Do NOT summarize or rewrite.
- Maintain original order.
- Include nested tables and multi-column text.
Return raw text only.
"""

            response = model.generate_content([uploaded_file, prompt])
            return response.text or ""

        except Exception as e:
            print(f"❌ Gemini PDF parsing failed for {file.name}: {e}")
            return ""

    else:
        return ""

def parse_resumes(
    resume_path: Path,
    output_excel_path: str = "candidate_feature_store.xlsx"
):

    files = sorted([
        f for f in resume_path.iterdir()
        if f.suffix.lower() in [".pdf", ".docx"]
    ])

    print(f"🚀 Found {len(files)} resumes")
    education_rows = []
    work_rows = []
    tech_rows = []
    demo_rows = []

    for file in tqdm(files):
        try:
            print(f"📄 Processing {file.name}")
            text = extract_text(file)
            features: FeatureProfile = llm_feature_engineering(text)

            # ================= EDUCATION SHEET =================
            for job in features.work.work_history:
                work_rows.append({
                    "candidate_name": features.name,
                    "company": job.company,
                    "job_title": job.job_title,
                    "years": job.years,
                    "sector": job.sector,
                    "strategy": job.investment_strategy,
                    "is_competitor": job.is_competitor,
                    "summary": job.summary,
                    "recency_score": job.recency_score,
                    # --- Candidate-level summary metrics kept in this sheet ---
                    "dominant_side": features.work.dominant_side,
                    "buyside_years_last_5": features.work.buyside_years_last_5,
                    "sellside_years_last_5": features.work.sellside_years_last_5,
                    "file_name": file.name
                })

            # --- UNPACK EDUCATION ---
            for edu in features.education.education_history:
                education_rows.append({
                    "candidate_name": features.name,
                    "university": edu.university,
                    "degree": edu.degree,
                    "major": edu.major,
                    "gpa": edu.gpa,
                    "graduation_year": edu.graduation_year,
                    "summary": edu.summary,
                    "recency_score": edu.recency_score,
                    # --- Candidate-level summary metrics ---
                    "highest_degree": features.education.highest_degree,
                    "institution_tier": features.education.institution_tier,
                    "file_name": file.name
                })

            # ================= TECHNICAL SHEET =================
            # for skill, score in features.technical.normalized_skills.items():
            skill_columns = build_skill_columns(features.technical)

            tech_rows.append({
                "file_name": file.name,
                "name": features.name,
                "additional_skills": features.technical.additional_skills,
                **skill_columns
            })

            # ================= DEMOGRAPHICS SHEET =================
            demo_rows.append({
                "file_name": file.name,
                "name": features.name,
                "total_years_experience": features.demographics.total_years_experience,
                "primary_geography": features.demographics.primary_geography,
                "extra-curriculars": features.demographics.extracurriculars
            })

        except Exception as e:
            print(f"❌ Failed on {file.name}: {e}")

    # ---------- Build DataFrames ----------
    edu_df = pd.DataFrame(education_rows)
    work_df = pd.DataFrame(work_rows)
    tech_df = pd.DataFrame(tech_rows)
    demo_df = pd.DataFrame(demo_rows)

    # ---------- Write Excel ----------
    with pd.ExcelWriter(output_excel_path, engine="xlsxwriter") as writer:
        edu_df.to_excel(writer, sheet_name="education", index=False)
        work_df.to_excel(writer, sheet_name="work_experience", index=False)
        tech_df.to_excel(writer, sheet_name="technical_skills", index=False)
        demo_df.to_excel(writer, sheet_name="demographics", index=False)

    print(f"\n✅ Feature Store Created → {output_excel_path}")

# parse_resumes(
#     resume_path=RESUME_PATH,
#     output_excel_path="candidate_feature_store.xlsx"
# )

def build_candidate_embedding_text(
    candidate_name: str,
    tech_row: dict,
    demo_row: dict,
    work_rows: list,
    edu_rows: list
) -> str:
    """
    Builds semantically rich embedding text from structured features.
    Raw resume text is intentionally NOT used.
    """

    lines = [
        f"Candidate Name: {candidate_name}",
        f"Total Experience: {demo_row.get('total_years_experience', 'NA')} years",
        f"Primary Geography: {demo_row.get('primary_geography', 'NA')}",
        "",
        "Technical Skills:"
    ]

    # ---- Skills ----
    for skill in SKILL_FIELDS:
        prefix = skill.lower()
        score = tech_row.get(f"{prefix}_skill_score", 0)

        if score > 0:
            lines.append(
                f"- {skill}: "
                f"score={round(score,3)}, "
                f"years_used={tech_row.get(f'{prefix}_years_used',0)}, "
                f"projects={tech_row.get(f'{prefix}_project_count',0)}"
            )

    lines.append("")
    lines.append("Recent Work Experience:")

    for w in work_rows[:3]:
        lines.append(
            f"- {w.get('company')} | {w.get('job_title')} | "
            f"{w.get('sector')} | {w.get('summary','')[:150]}"
        )

    lines.append("")
    lines.append("Education:")

    for e in edu_rows[:2]:
        lines.append(
            f"- {e.get('university')} | {e.get('degree')} | {e.get('major')}"
        )

    return "\n".join(lines)

def build_and_store_feature_embeddings_streaming(excel_path):

    wb = load_workbook(excel_path, read_only=True)
    tech_ws = wb["technical_skills"]
    demo_ws = wb["demographics"]
    work_ws = wb["work_experience"]
    edu_ws  = wb["education"]

    # ---- Build lightweight lookup maps ----
    demo_map = {}
    for row in demo_ws.iter_rows(min_row=2, values_only=True):
        name = row[1]
        demo_map[name] = {
            "total_years_experience": row[2],
            "primary_geography": row[3],
            "extra-curriculars": row[4],
        }

    work_map = {}
    for row in work_ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        work_map.setdefault(name, []).append({
            "company": row[1],
            "job_title": row[2],
            "sector": row[4],
            "summary": row[7]
        })

    edu_map = {}
    for row in edu_ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        edu_map.setdefault(name, []).append({
            "university": row[1],
            "degree": row[2],
            "major": row[3]
        })

    dim = embedder.get_sentence_embedding_dimension()
    index = faiss.IndexFlatIP(dim)

    embedding_texts = []
    metadata = []

    for row in tech_ws.iter_rows(min_row=2, values_only=True):

        file_name, name, additional_skills, *scores = row

        tech_row = {
            f"{s.lower()}_skill_score": scores[i]
            for i, s in enumerate(SKILL_FIELDS)
        }

        text = build_candidate_embedding_text(
            candidate_name=name,
            tech_row=tech_row,
            demo_row=demo_map.get(name, {}),
            work_rows=work_map.get(name, [])[:3],
            edu_rows=edu_map.get(name, [])[:2]
        )

        vec = embedder.encode([text], normalize_embeddings=True).astype("float32")
        index.add(vec)

        embedding_texts.append(text)
        metadata.append({"name": name})

        del vec, text, tech_row
        gc.collect()

    faiss.write_index(index, str(FAISS_INDEX_PATH))

    with open(METADATA_PATH, "w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2)

    with open(EMBED_TEXT_PATH, "w", encoding="utf-8") as f:
        json.dump(embedding_texts, f, indent=2)

    print("✅ Feature embeddings stored successfully")


CONTINENTS = [
    "Asia",
    "Europe",
    "North America",
    "South America",
    "Africa",
    "Oceania"
]

st.set_page_config(layout="wide")

llm_client = OpenAI(
    api_key=GEMINI_API_KEY,
    base_url=GEMINI_BASE_URL
)

cc = coco.CountryConverter()
JD_MIN_SCORE = 0.45

@lru_cache(maxsize=4096)
def normalize_to_continent(location_text: str) -> str:
    """
    Maps free-text location → continent using fast keyword rules first,
    then falls back to country_converter.
    """

    if not isinstance(location_text, str) or not location_text.strip():
        return "Unknown"

    lower = location_text.strip().lower()
    
    us_keywords = [
        "new york", "nyc", "san francisco", "sf", "boston", "ma",
        "chicago", "seattle", "austin", "texas", "california",
        "usa", "united states", "u.s.", "miami", "ny"
    ]

    europe_keywords = [
        "london", "paris", "berlin", "amsterdam", "zurich",
        "geneva", "frankfurt", "milan", "madrid", "uk", "germany",
        "france", "switzerland", "netherlands", "italy", "spain"
    ]

    if any(k in lower for k in us_keywords):
        return "North America"

    if any(k in lower for k in europe_keywords):
        return "Europe"
    try:
        country = cc.convert(names=[location_text], to="name_short", not_found=None)
        continent = cc.convert(names=[country], to="continent", not_found=None)

        if isinstance(continent, list):
            continent = continent[0]

        if continent in CONTINENTS:
            return continent
    except Exception as e:
        logger.debug(f"country_converter failed for '{location_text}': {e}")

    return "Unknown"


@st.cache_data(show_spinner=False)
def load_data():

    xls = pd.ExcelFile(EXCEL_PATH)

    edu = pd.read_excel(xls, "education")
    work = pd.read_excel(xls, "work_experience")
    tech = pd.read_excel(xls, "technical_skills")
    demo = pd.read_excel(xls, "demographics")

    sector_map = (
        work.groupby("candidate_name")["sector"]
        .apply(lambda x: ", ".join(sorted(set([str(v) for v in x if pd.notna(v)]))))
        .reset_index()
        .rename(columns={"candidate_name": "name", "sector": "sector_experience"})
    )

    strategy_map = (
        work.groupby("candidate_name")["strategy"]
        .apply(lambda x: ", ".join(sorted(set([str(v) for v in x if pd.notna(v)]))))
        .reset_index()
        .rename(columns={"candidate_name": "name", "strategy": "strategy_experience"})
    )

    edu_summary = (
        edu.groupby("candidate_name")
        .agg(
            highest_degree=("highest_degree", "first"),
            institution_tier=("institution_tier", "first"),
            gpa=("gpa", "first"),
            graduation_year=("graduation_year", "first"),
        )
        .reset_index()
        .rename(columns={"candidate_name": "name"})
    )

    competitor_map = (
        work.groupby("candidate_name")["is_competitor"]
        .max()
        .reset_index()
        .rename(columns={"candidate_name": "name", "is_competitor": "has_competitor_exp"})
    )

    skill_score_cols = [c for c in tech.columns if c.endswith("_skill_score")]

    def build_skill_map(row):
        skill_map = {}
        for col in skill_score_cols:
            val = row.get(col, 0)
            if pd.notna(val) and float(val) > 0:
                clean = col.replace("_skill_score", "").replace("_", " ")
                skill_map[clean] = round(float(val), 3)
        return skill_map

    tech["skill_map"] = tech.apply(build_skill_map, axis=1)
    tech_small = tech[["file_name", "name", "skill_map"]]

    master = (
        demo.merge(tech_small, on=["file_name", "name"], how="left")
            .merge(sector_map, on="name", how="left")
            .merge(strategy_map, on="name", how="left")
            .merge(edu_summary, on="name", how="left")
    )

    master["skill_map"] = master["skill_map"].fillna({})
    master["sector_experience"] = master["sector_experience"].fillna("")
    master["strategy_experience"] = master["strategy_experience"].fillna("")
    
    master = master.merge(competitor_map, on="name", how="left")
    master["has_competitor_exp"] = master["has_competitor_exp"].fillna(False)

    return master, work


master_df, work_df = load_data()
# ================= ACADEMIC BENCHMARKS ================

def build_academic_benchmarks(df):

    benchmarks = {}

    # ---- GPA percentile ----
    if "gpa" in df.columns:
        gpas = pd.to_numeric(df["gpa"], errors="coerce").dropna()
        benchmarks["gpa_p75"] = gpas.quantile(0.75) if not gpas.empty else 0.75
    else:
        logger.warning("gpa column missing — using default GPA benchmark")
        benchmarks["gpa_p75"] = 0.75

    # ---- Graduation recency percentile ----
    if "graduation_year" in df.columns:
        current_year = datetime.now().year
        grad_years = pd.to_numeric(df["graduation_year"], errors="coerce").dropna()
        recency = current_year - grad_years
        benchmarks["recent_grad_p25"] = recency.quantile(0.25) if not recency.empty else 8
    else:
        logger.warning("graduation_year column missing — using default recency benchmark")
        benchmarks["recent_grad_p25"] = 8

    # ---- Tier distribution ----
    if "institution_tier" in df.columns:
        tier_counts = (
            df["institution_tier"]
            .fillna("")
            .str.lower()
            .value_counts(normalize=True)
        )
        benchmarks["tier1_share"] = tier_counts.get("tier-1", 0)
    else:
        logger.warning("institution_tier column missing — defaulting tier share to 0")
        benchmarks["tier1_share"] = 0

    return benchmarks



ACADEMIC_BENCHMARKS = build_academic_benchmarks(master_df)
ALL_SKILLS = sorted({k for sm in master_df["skill_map"] for k in sm.keys()})

# =====================================================
# ================= UTILITIES =========================
# =====================================================

from datetime import datetime

def is_strong_academics(candidate):
    """
    Population-normalized academic strength signal.
    Candidate must outperform population on >=2 dimensions.
    """

    score = 0
    current_year = datetime.now().year
    tier = str(candidate.get("institution_tier", "")).lower()
    if "tier-1" in tier:
        score += 1
    try:
        gpa = float(candidate.get("gpa"))
        if gpa >= ACADEMIC_BENCHMARKS["gpa_p75"]:
            score += 1
    except Exception:
        pass
    try:
        grad_year = int(candidate.get("graduation_year"))
        years_since_grad = current_year - grad_year
        if years_since_grad <= ACADEMIC_BENCHMARKS["recent_grad_p25"]:
            score += 1
    except Exception:
        pass
    return score >= 2


def clean_llm_json(raw_text: str) -> str:
    text = raw_text.strip()
    text = re.sub(r"^```[a-zA-Z]*", "", text)
    text = re.sub(r"```$", "", text).strip()
    match = re.search(r"\{.*\}", text, flags=re.DOTALL)
    if match:
        text = match.group(0)
    return text


def render_docx_readable(file_path: Path) -> str:
    doc = Document(file_path)
    lines = []

    for p in doc.paragraphs:
        text = p.text.strip()
        if not text:
            continue

        if p.style and "List" in p.style.name:
            lines.append(f"• {text}")
            continue

        is_heading = False
        if p.runs:
            bold_runs = sum(1 for r in p.runs if r.bold)
            if bold_runs >= max(1, len(p.runs) // 2):
                is_heading = True

        if text.isupper() or is_heading:
            lines.append(f"\n### {text}\n")
        else:
            lines.append(text)

    return "\n".join(lines)

@st.cache_data(show_spinner=False, ttl=1800)
def extract_job_intent(job_text: str):

    prompt = f"""
You are a quantitative hiring analyst.

Extract structured hiring intent from this job description.

Return JSON with fields:
- target_strategy: one of ["systematic", "fundamental", "quantitative", "any"]
- core_skills: list of critical skills
- secondary_skills: list of nice-to-have skills
- target_experience: integer (years or null)
- education_importance: one of ["low", "medium", "high"]

Job Description:
{job_text}

Output JSON only.
"""

    response = llm_client.chat.completions.create(
        model=MODEL_NAME,
        temperature=0,
        messages=[{"role": "user", "content": prompt}],
    )

    raw = response.choices[0].message.content
    cleaned = clean_llm_json(raw)
    return json.loads(cleaned)

def apply_hard_filters(
    df,
    strategy,
    min_exp,
    max_exp,
    required_skills,
    selected_continents
):
    rows = []

    for _, r in df.iterrows():
        if strategy != "All":
            if strategy.lower() not in str(r["strategy_experience"]).lower():
                continue

        exp = r.get("total_years_experience", 0) or 0
        if not (min_exp <= exp <= max_exp):
            continue
        if required_skills:
            skill_keys = " ".join([k.lower() for k in r.get("skill_map", {}).keys()])
            if not any(s.lower() in skill_keys for s in required_skills):
                continue
        if selected_continents:
            candidate_continent = normalize_to_continent(
                str(r.get("primary_geography", ""))
            )
            if candidate_continent not in selected_continents:
                continue

        rows.append(r)

    return pd.DataFrame(rows)

def score_candidate(candidate, intent):

    if not intent:
        return 0.5

    strategy_score = 1.0
    if intent["target_strategy"] != "any":
        if intent["target_strategy"] in str(candidate["strategy_experience"]).lower():
            strategy_score = 1.0
        else:
            strategy_score = 0.3

    skill_map = candidate.get("skill_map", {})

    def avg_skill(skill_list):
        vals = []
        for s in skill_list:
            for k, v in skill_map.items():
                if s.lower() in k.lower():
                    vals.append(v)
        return sum(vals) / len(vals) if vals else 0.0

    core_skill_score = avg_skill(intent.get("core_skills", []))
    secondary_skill_score = avg_skill(intent.get("secondary_skills", []))
    skill_score_final = 0.7 * core_skill_score + 0.3 * secondary_skill_score

    tier = str(candidate.get("institution_tier", "")).lower()
    edu_base = 1.0 if "tier-1" in tier else 0.6 if "tier-2" in tier else 0.3
    edu_weight = {"low": 0.4, "medium": 0.7, "high": 1.0}.get(
        intent.get("education_importance", "medium"), 0.7
    )
    education_score = edu_base * edu_weight

    candidate_exp = candidate.get("total_years_experience", 0) or 0
    target_exp = intent.get("target_experience") or candidate_exp

    exp_gap = target_exp - candidate_exp
    if exp_gap <= 0:
        experience_score = 1.0
    else:
        experience_score = max(0.3, 1 - exp_gap / 5)

    experience_score += 0.3 * skill_score_final
    experience_score += 0.2 * education_score
    experience_score = min(experience_score, 1.0)

    final_score = (
        0.30 * strategy_score +
        0.30 * skill_score_final +
        0.20 * experience_score +
        0.20 * education_score
    )

    return round(final_score, 4)


def rank_candidates(df, intent):

    if df.empty:
        return pd.DataFrame()

    jd_text = st.session_state.get("raw_jd_text", "")

    require_competitor = any(
        kw in jd_text for kw in ["competitor", "tier-1 firm", "top hedge fund", "bulge bracket"]
    )

    require_top_tier_edu = any(
        kw in jd_text for kw in ["top tier", "tier-1", "ivy", "iit", "elite university", "strong academics"]
    )

    scored_rows = []

    for _, r in df.iterrows():
        r = r.copy()
        r["final_score"] = score_candidate(r, intent)

        if require_competitor:
            if not bool(r.get("has_competitor_exp", False)):
                continue

        if require_top_tier_edu:
            if not is_strong_academics(r):
                continue


        if intent and r["final_score"] < JD_MIN_SCORE:
            continue

        scored_rows.append(r)

    if not scored_rows:
        return pd.DataFrame()

    return (
        pd.DataFrame(scored_rows)
          .sort_values("final_score", ascending=False)
    )


def open_resume(file_name: str):
    st.session_state.selected_candidate = file_name

st.sidebar.title("🎯 Hiring Inputs")

job_text = st.sidebar.text_area(
    "Job Description / Hiring Intent",
    placeholder="Systematic researcher, Python heavy, SQL helpful, 3–5 years, strong academics"
)

search_clicked = st.sidebar.button("🔍 Search Job Intent")

st.sidebar.markdown("### 🔒 Hard Filters")

strategy = st.sidebar.selectbox(
    "Investment Strategy",
    ["All", "Systematic", "Fundamental", "Quantitative"]
)

min_exp, max_exp = st.sidebar.slider(
    "Years of Experience",
    0, 25, (0, 25)
)

required_skills = st.sidebar.multiselect(
    "Required Skills",
    ALL_SKILLS
)

selected_continents = st.sidebar.multiselect(
    "Preferred Continent",
    CONTINENTS
)

st.title("🏦 Trading Talent Intelligence Platform")

if "selected_candidate" not in st.session_state:
    st.session_state.selected_candidate = None

if st.session_state.selected_candidate is not None:

    selected_file = st.session_state.selected_candidate
    row_match = master_df[master_df["file_name"] == selected_file]

    if row_match.empty:
        st.error("Candidate not found.")
        st.session_state.selected_candidate = None
        st.stop()

    row = row_match.iloc[0]

    if st.button("⬅ Back to Results"):
        st.session_state.selected_candidate = None
        st.rerun()

    st.subheader(f"📄 Candidate Profile — {row.get('name','')}")

    resume_path = RESUME_FOLDER / selected_file

    if not resume_path.exists():
        st.error("Resume file not found.")
        st.stop()

    if resume_path.suffix.lower() == ".pdf":
        pdf_bytes = resume_path.read_bytes()
        base64_pdf = base64.b64encode(pdf_bytes).decode("utf-8")

        st.markdown(
            f"<iframe src='data:application/pdf;base64,{base64_pdf}' width='100%' height='800px'></iframe>",
            unsafe_allow_html=True
        )

    elif resume_path.suffix.lower() == ".docx":
        try:
            formatted_text = render_docx_readable(resume_path)
            st.markdown(formatted_text)
        except Exception as e:
            st.warning(f"DOCX preview failed: {e}")
            st.download_button("⬇️ Download Resume", resume_path.read_bytes())

    else:
        st.warning("Unsupported resume format.")

    st.stop()

if "intent" not in st.session_state:
    st.session_state.intent = {}

if "raw_jd_text" not in st.session_state:
    st.session_state.raw_jd_text = ""

if search_clicked and job_text.strip():
    try:
        st.session_state.intent = extract_job_intent(job_text)
        st.session_state.raw_jd_text = job_text.lower()
    except Exception:
        st.warning("⚠️ Failed to parse job intent")

intent = st.session_state.intent

filtered = apply_hard_filters(
    master_df,
    strategy,
    min_exp,
    max_exp,
    required_skills,
    selected_continents
)

ranked_df = rank_candidates(filtered, intent)
if intent and ranked_df.empty:
    st.warning("⚠️ Job description is too restrictive — no candidates matched strongly.")

st.info(f"🔢 {len(ranked_df)} candidates matched")

st.markdown("## 📊 Talent Overview")

if ranked_df.empty:
    st.warning("⚠️ No candidates match the current filters.")
else:
    col1, col2, col3 = st.columns(3)

    with col1:
        sector_counts = (
            work_df[work_df["candidate_name"].isin(ranked_df["name"])]
            ["sector"]
            .value_counts()
        )
        if not sector_counts.empty:
            st.markdown("### 🏭 Sector Distribution")
            st.pyplot(sector_counts.plot.pie(autopct="%1.0f%%").figure)

    with col2:
        skill_accumulator = {}
        for smap in ranked_df["skill_map"]:
            for skill, score in smap.items():
                skill_accumulator.setdefault(skill, []).append(score)

        if skill_accumulator:
            avg_skill_scores = {
                k: round(sum(v) / len(v), 3)
                for k, v in skill_accumulator.items()
            }
            st.markdown("### 🧠 Average Skill Strength")
            st.bar_chart(
                pd.Series(avg_skill_scores)
                  .sort_values(ascending=False)
                  .head(12)
            )

    with col3:
        def classify_strategy(s):
            s = str(s).lower()
            if "systematic" in s or "quant" in s:
                return "Systematic / Quant"
            return "Fundamental / Other"

        split_counts = ranked_df["strategy_experience"].apply(classify_strategy).value_counts()
        if not split_counts.empty:
            st.markdown("### ⚖️ Strategy Split")
            st.bar_chart(split_counts)

cols = st.columns(3)

for idx, row in ranked_df.iterrows():
    col = cols[idx % 3]

    with col:
        st.markdown(f"### 👤 {row.get('name','')}")
        st.write(f"🌍 Location: {row.get('primary_geography','')}")
        st.write(f"📊 Experience: {row.get('total_years_experience','')} yrs")
        st.write(f"🎯 Strategies: {row.get('strategy_experience','')}")
        st.write(f"🏭 Sectors: {row.get('sector_experience','')}")

        btn_key = f"view_{row['file_name']}_{idx}"

        st.button(
            "View Resume",
            key=btn_key,
            on_click=open_resume,
            args=(row["file_name"],)
        )
